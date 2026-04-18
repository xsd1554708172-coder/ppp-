# -*- coding: utf-8 -*-
from __future__ import annotations
"""Minimal refresh script for current official V3 runtime files.
Boundary: MINIMAL REFRESH ONLY, not a full package rebuild.
"""
import json
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

THIS_FILE = Path(__file__).resolve()
SCRIPT_DIR = THIS_FILE.parent
PACKAGE_ROOT = SCRIPT_DIR.parents[1]  # .../full
ACTIVE_RESULTS_DIR = PACKAGE_ROOT / '01_第3到第8部分_最终修正版' / '第5部分_核心实证识别' / '（5.1）识别框架、并表与模型设定'
ACTIVE_RISK_DIR = PACKAGE_ROOT / '02_第9部分_最终可用内容' / '（补充）项目级建模底表'
ACTIVE_CONFIG = PACKAGE_ROOT / '02_第9部分_最终可用内容' / 'config.json'

def find_latest(directory: Path, prefix: str, suffix: str) -> Path:
    matches = sorted([p for p in directory.glob(f'{prefix}*{suffix}') if p.is_file()])
    if not matches:
        raise FileNotFoundError(f'未在 {directory} 中找到以前缀 {prefix} 和后缀 {suffix} 匹配的文件')
    return matches[-1]

def sheet_to_df(xlsx_path: Path, sheet_name: str) -> pd.DataFrame:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    return pd.DataFrame(rows[1:], columns=rows[0])

def main() -> None:
    panel_xlsx = find_latest(ACTIVE_RESULTS_DIR, 'PPP_3.6_model_ready_panel_v3_统一口径_方案2_实际执行版_', '.xlsx')
    risk_xlsx = find_latest(ACTIVE_RISK_DIR, 'PPP_project_level_risk_model_data_v3_无泄漏严格版_标签统一_', '.xlsx')
    panel_csv = panel_xlsx.with_suffix('.csv')
    risk_csv = risk_xlsx.with_suffix('.csv')
    panel_df = sheet_to_df(panel_xlsx, 'panel_v3')
    panel_df.to_csv(panel_csv, index=False, encoding='utf-8-sig')
    risk_df = sheet_to_df(risk_xlsx, 'full_data_10000')
    risk_df.to_csv(risk_csv, index=False, encoding='utf-8-sig')
    cfg = json.loads(ACTIVE_CONFIG.read_text(encoding='utf-8'))
    cfg['data_path'] = f'（补充）项目级建模底表/{risk_csv.name}'
    ACTIVE_CONFIG.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding='utf-8')
    print('OK')
    print(panel_xlsx)
    print(risk_xlsx)

if __name__ == '__main__':
    main()
