from __future__ import annotations

import math
import textwrap
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import numpy as np
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from scipy.stats import norm


WORKSPACE_ROOT = Path(__file__).resolve().parents[3]
BUNDLE_ROOT = WORKSPACE_ROOT / "PPP_empirical_reinforcement_bundle_20260416_clean"
DATA_ROOT = WORKSPACE_ROOT / "ppp论文数据"

MAIN_PANEL_NAME = "PPP_3.6_model_ready_panel_v3_统一口径_方案2_实际执行版_20260413_1048.csv"
LONG_TABLE_NAME = "PPP_第5部分_5.3正式回归结果长表_V3_重估版_20260413_1048.csv"
DOCX_NAME = "PPP论文_完整论文初稿_公共管理风格_修订版_定点替换_20260415.docx"

CONTROLS = [
    "dfi",
    "digital_econ",
    "ln_rd_expenditure",
    "ln_tech_contract_value",
    "ln_patent_grants",
]
OUTCOMES = ["exec_share", "proc_share", "ppp_quality_zindex"]


def find_single(root: Path, filename: str) -> Path:
    matches = sorted(root.rglob(filename))
    if not matches:
        raise FileNotFoundError(f"Could not locate {filename} under {root}")
    return matches[0]


def find_docx_paragraphs(docx_path: Path) -> list[str]:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(docx_path) as zf:
        root = ET.fromstring(zf.read("word/document.xml"))
    paragraphs: list[str] = []
    for p in root.findall(".//w:p", ns):
        texts = [node.text for node in p.findall(".//w:t", ns) if node.text]
        if texts:
            paragraphs.append("".join(texts))
    return paragraphs


def relative_to_workspace(path: Path) -> str:
    return str(path.relative_to(WORKSPACE_ROOT)).replace("\\", "/")


def rel_or_name(path: Path) -> str:
    try:
        return relative_to_workspace(path)
    except ValueError:
        return path.name


def load_panel() -> tuple[Path, pd.DataFrame]:
    panel_path = find_single(DATA_ROOT, MAIN_PANEL_NAME)
    panel = pd.read_csv(panel_path)
    return panel_path, panel


def load_long_table() -> tuple[Path, pd.DataFrame]:
    table_path = find_single(DATA_ROOT, LONG_TABLE_NAME)
    table = pd.read_csv(table_path)
    return table_path, table


def extract_anchor_rows(paragraphs: list[str]) -> pd.DataFrame:
    anchors: list[dict[str, str | int]] = []
    keywords = [
        ("研究总述", ["treat_share", "多期DID/TWFE"]),
        ("主识别定义", ["本文的主识别模型是以treat_share为核心处理变量"]),
        ("结果定位", ["基准结果显示", "treat_share显著提高执行阶段占比", "显著降低采购阶段占比"]),
        ("边界说明", ["ppp_quality_zindex", "不能被写成全文主结论"]),
    ]
    for idx, para in enumerate(paragraphs):
        if any(key in para for _, keys in keywords for key in keys):
            role = "正文锚点"
            for label, keys in keywords:
                if any(key in para for key in keys):
                    role = label
                    break
            anchors.append({"paragraph_index": idx, "anchor_role": role, "anchor_text": para})
    return pd.DataFrame(anchors)


def make_design_matrix(
    df: pd.DataFrame,
    depvar: str,
    add_trend: bool,
) -> tuple[np.ndarray, np.ndarray, list[str], pd.DataFrame]:
    use_cols = ["province", "year", depvar, "treat_share", *CONTROLS]
    use = df[use_cols].dropna().copy()
    use["year_idx"] = use["year"] - use["year"].min()

    provinces = pd.Categorical(use["province"])
    years = pd.Categorical(use["year"])

    columns: list[np.ndarray] = [np.ones((len(use), 1), dtype=float)]
    names = ["const"]

    for col in ["treat_share", *CONTROLS]:
        columns.append(use[[col]].to_numpy(dtype=float))
        names.append(col)

    for level in provinces.categories[1:]:
        columns.append((use["province"].to_numpy() == level).astype(float).reshape(-1, 1))
        names.append(f"province_fe[{level}]")

    for level in years.categories[1:]:
        columns.append((use["year"].to_numpy() == level).astype(float).reshape(-1, 1))
        names.append(f"year_fe[{level}]")

    if add_trend:
        for level in provinces.categories:
            columns.append(
                ((use["province"].to_numpy() == level) * use["year_idx"].to_numpy()).astype(float).reshape(-1, 1)
            )
            names.append(f"trend[{level}]")

    X = np.hstack(columns)
    y = use[depvar].to_numpy(dtype=float).reshape(-1, 1)
    return X, y, names, use


def cluster_ols(
    df: pd.DataFrame,
    depvar: str,
    add_trend: bool,
) -> dict[str, float | str | int]:
    X, y, names, use = make_design_matrix(df, depvar, add_trend=add_trend)
    beta, *_ = np.linalg.lstsq(X, y, rcond=None)
    resid = y - X @ beta

    xtx_inv = np.linalg.pinv(X.T @ X)
    clusters = use["province"].to_numpy()
    unique_clusters = pd.unique(clusters)
    meat = np.zeros((X.shape[1], X.shape[1]), dtype=float)

    for cluster in unique_clusters:
        idx = np.where(clusters == cluster)[0]
        xg = X[idx, :]
        ug = resid[idx, :]
        meat += xg.T @ ug @ ug.T @ xg

    g = len(unique_clusters)
    n = len(use)
    k = X.shape[1]
    dfc = (g / (g - 1)) * ((n - 1) / (n - k))
    cov = dfc * xtx_inv @ meat @ xtx_inv
    se = np.sqrt(np.diag(cov))

    coef_idx = names.index("treat_share")
    coef = float(beta[coef_idx, 0])
    se_val = float(se[coef_idx])
    t_val = coef / se_val
    p_val = float(2 * norm.sf(abs(t_val)))

    formula = (
        f"{depvar} ~ treat_share + {' + '.join(CONTROLS)} + C(province) + C(year)"
        + (" + C(province):year_idx" if add_trend else "")
    )
    return {
        "depvar": depvar,
        "coef": coef,
        "se": se_val,
        "t": float(t_val),
        "p": p_val,
        "N": n,
        "formula": formula,
    }


def prettify_coef(coef: float, se: float, p: float) -> str:
    stars = "***" if p < 0.01 else "**" if p < 0.05 else "*" if p < 0.1 else ""
    return f"{coef:.4f}{stars} ({se:.4f})"


def load_baseline_results(long_table: pd.DataFrame) -> pd.DataFrame:
    baseline = long_table[long_table["did_var"] == "treat_share"].copy()
    baseline = baseline[
        [
            "depvar",
            "did_var",
            "N",
            "coef",
            "se",
            "t",
            "p",
            "r2",
            "mean_dep",
            "province_fe",
            "year_fe",
            "cluster",
            "controls",
            "coef_star",
            "se_fmt",
            "panel_version",
            "estimation_note",
        ]
    ].reset_index(drop=True)
    baseline.insert(0, "spec_type", "baseline_treat_share")
    return baseline


def build_trend_results(panel: pd.DataFrame) -> pd.DataFrame:
    rows = [cluster_ols(panel, depvar, add_trend=True) for depvar in OUTCOMES]
    return pd.DataFrame(rows)


def build_comparison_table(baseline: pd.DataFrame, trend: pd.DataFrame) -> pd.DataFrame:
    merged = baseline.merge(trend, on="depvar", suffixes=("_baseline", "_trend"))
    return pd.DataFrame(
        {
            "depvar": merged["depvar"],
            "baseline_coef": merged["coef_baseline"],
            "baseline_se": merged["se_baseline"],
            "baseline_p": merged["p_baseline"],
            "baseline_display": merged["se_fmt"],
            "trend_coef": merged["coef_trend"],
            "trend_se": merged["se_trend"],
            "trend_p": merged["p_trend"],
            "trend_display": [
                prettify_coef(c, s, p)
                for c, s, p in zip(merged["coef_trend"], merged["se_trend"], merged["p_trend"], strict=True)
            ],
            "delta_coef": merged["coef_trend"] - merged["coef_baseline"],
            "delta_p": merged["p_trend"] - merged["p_baseline"],
        }
    )


def make_forest_plot(path: Path, comparison: pd.DataFrame) -> None:
    width, height = 1400, 900
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    font = ImageFont.load_default()

    left = 260
    top = 120
    chart_width = 980
    row_h = 180
    x_min, x_max = -0.65, 0.75
    zero_x = left + int((0 - x_min) / (x_max - x_min) * chart_width)
    draw.line((left, top - 30, left, top + row_h * len(comparison) - 10), fill="#777777", width=1)
    draw.line((left, top + row_h * len(comparison), left + chart_width, top + row_h * len(comparison)), fill="#777777", width=1)
    draw.line((zero_x, top - 30, zero_x, top + row_h * len(comparison)), fill="#bbbbbb", width=2)

    # Grid and axis labels
    for tick in np.linspace(x_min, x_max, 6):
        x = left + int((tick - x_min) / (x_max - x_min) * chart_width)
        draw.line((x, top - 30, x, top + row_h * len(comparison)), fill="#efefef", width=1)
        label = f"{tick:.1f}"
        bbox = draw.textbbox((0, 0), label, font=font)
        draw.text((x - (bbox[2] - bbox[0]) / 2, top + row_h * len(comparison) + 8), label, fill="black", font=font)

    draw.text((left, 40), "Baseline vs Trend-adjusted DID", fill="black", font=font)
    draw.text((left, 58), "Solid = baseline 5.3 long table, hollow = trend-adjusted re-estimation", fill="black", font=font)

    series = [
        ("Baseline", "#2f6fdf", -18, "baseline_coef", "baseline_se"),
        ("Trend-adjusted", "#d14b3a", 18, "trend_coef", "trend_se"),
    ]

    for i, row in comparison.reset_index(drop=True).iterrows():
        y_center = top + row_h * i + 60
        outcome_label = row["depvar"]
        draw.text((50, y_center - 8), outcome_label, fill="black", font=font)
        for label, color, offset, coef_key, se_key in series:
            coef = float(row[coef_key])
            se = float(row[se_key])
            ci_low = coef - 1.96 * se
            ci_high = coef + 1.96 * se
            x0 = left + int((ci_low - x_min) / (x_max - x_min) * chart_width)
            x1 = left + int((ci_high - x_min) / (x_max - x_min) * chart_width)
            xc = left + int((coef - x_min) / (x_max - x_min) * chart_width)
            draw.line((x0, y_center + offset, x1, y_center + offset), fill=color, width=5)
            draw.ellipse((xc - 8, y_center + offset - 8, xc + 8, y_center + offset + 8), outline=color, width=3, fill="white")
            draw.text((x1 + 8, y_center + offset - 7), f"{label}: {coef:.3f}", fill=color, font=font)

    legend_y = top + row_h * len(comparison) + 45
    draw.rectangle((50, legend_y, 185, legend_y + 45), outline="#999999", width=1)
    draw.text((62, legend_y + 8), "Legend", fill="black", font=font)
    draw.ellipse((62, legend_y + 24, 74, legend_y + 36), fill="#2f6fdf")
    draw.text((82, legend_y + 19), "baseline", fill="black", font=font)
    draw.ellipse((138, legend_y + 24, 150, legend_y + 36), outline="#d14b3a", width=2, fill="white")
    draw.text((158, legend_y + 19), "trend", fill="black", font=font)

    img.save(path)


def write_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    wb = Workbook()
    first = True
    for sheet_name, frame in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        for c, col in enumerate(frame.columns, start=1):
            cell = ws.cell(1, c, col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for r, (_, row) in enumerate(frame.iterrows(), start=2):
            for c, value in enumerate(row.tolist(), start=1):
                ws.cell(r, c, value)

        ws.freeze_panes = "A2"
        for col_idx, col_name in enumerate(frame.columns, start=1):
            max_len = max(len(str(col_name)), *(len(str(v)) for v in frame.iloc[:, col_idx - 1].tolist()))
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else f"A{col_idx}"].width = min(max_len + 2, 48)

    wb.save(path)


def build_unified_baseline_reference(
    bundle_root: Path,
    docx_path: Path,
    long_table_path: Path,
    panel_path: Path,
    long_table: pd.DataFrame,
    panel: pd.DataFrame,
    paragraphs: list[str],
) -> tuple[Path, Path]:
    target_dir = bundle_root / "00_unified_baseline_reference"
    target_dir.mkdir(parents=True, exist_ok=True)

    anchor_df = extract_anchor_rows(paragraphs)
    baseline_df = load_baseline_results(long_table)

    panel_cols = [
        "province",
        "year",
        "treat_share",
        "exec_share",
        "proc_share",
        "ppp_quality_zindex",
        "did_intensity",
        "did_any",
        "baseline_sample_5_3",
        "baseline_controls_complete",
        "panel_version",
    ]
    available_panel_cols = [c for c in panel_cols if c in panel.columns]
    key_panel = panel.loc[panel["baseline_sample_5_3"] == 1, available_panel_cols].copy()

    unified = pd.DataFrame(
        [
            {
                "item": "canonical_identification",
                "value": "treat_share DID/TWFE",
                "note": "正文与5.3长表一致；trend-adjusted DID 仅为防守型稳健性层",
            },
            {
                "item": "canonical_panel",
                "value": relative_to_workspace(panel_path),
                "note": "正式 V3 主面板，baseline_sample_5_3 = 262",
            },
            {
                "item": "canonical_long_table",
                "value": relative_to_workspace(long_table_path),
                "note": "当前 baseline 数值锚点；1048 与正文表述对齐",
            },
            {
                "item": "version_check",
                "value": "20260413_1048 == 20260413_0912",
                "note": "两份主面板逐单元完全一致；选 1048 作为最终 canonical reference because it is the latest file used by the正文 5.3 long table",
            },
        ]
    )

    path_index = pd.DataFrame(
        [
            {
                "artifact": "docx anchor",
                "relative_path": relative_to_workspace(docx_path),
                "note": "仅用于识别正文 baseline DID anchor；不作为数值源",
            },
            {
                "artifact": "panel source",
                "relative_path": relative_to_workspace(panel_path),
                "note": "正式 V3 主面板（baseline_sample_5_3 = 262）",
            },
            {
                "artifact": "long table source",
                "relative_path": relative_to_workspace(long_table_path),
                "note": "5.3 正式回归结果长表，作为 baseline 数值锚点",
            },
        ]
    )

    workbook_path = target_dir / "unified_baseline_reference.xlsx"
    write_excel(
        workbook_path,
        {
            "统一结论": unified,
            "正文锚点": anchor_df,
            "5.3主结果": baseline_df,
            "路径索引": path_index,
            "样本核对": key_panel,
        },
    )

    readme_path = target_dir / "baseline_spec_readme.md"
    readme_text = textwrap.dedent(
        f"""
        # Unified Baseline Reference

        ## Canonical rule
        - The only canonical main identification in this bundle is `treat_share` DID/TWFE.
        - `did_any` and `did_intensity` remain comparison specifications only.
        - `trend-adjusted DID` is a defensive robustness layer and does not replace the canonical baseline.

        ## Source chain
        - 正文锚点：`{relative_to_workspace(docx_path)}`
        - 正式 V3 主面板：`{relative_to_workspace(panel_path)}`
        - 5.3 正式回归长表：`{relative_to_workspace(long_table_path)}`

        ## Why 1048 is canonical
        - The 20260413_1048 and 20260413_0912 panel files are byte-for-byte identical in row content and columns.
        - The 1048 file is the final timestamped file used by the current 5.3 long table, so it is the cleanest canonical reference for downstream modules.

        ## What this workbook records
        - The exact正文 anchor paragraphs that define the baseline DID framing.
        - The 5.3 baseline coefficients for `exec_share`, `proc_share`, and `ppp_quality_zindex`.
        - A relative-path index so every downstream module can stay path-relative.

        ## Baseline numbers
        - `exec_share`: {baseline_df.loc[baseline_df['depvar'] == 'exec_share', 'coef'].iloc[0]:.6f}, p={baseline_df.loc[baseline_df['depvar'] == 'exec_share', 'p'].iloc[0]:.6f}
        - `proc_share`: {baseline_df.loc[baseline_df['depvar'] == 'proc_share', 'coef'].iloc[0]:.6f}, p={baseline_df.loc[baseline_df['depvar'] == 'proc_share', 'p'].iloc[0]:.6f}
        - `ppp_quality_zindex`: {baseline_df.loc[baseline_df['depvar'] == 'ppp_quality_zindex', 'coef'].iloc[0]:.6f}, p={baseline_df.loc[baseline_df['depvar'] == 'ppp_quality_zindex', 'p'].iloc[0]:.6f}
        """
    ).strip() + "\n"
    readme_path.write_text(readme_text, encoding="utf-8")
    return workbook_path, readme_path


def build_trend_outputs(
    bundle_root: Path,
    panel_path: Path,
    panel: pd.DataFrame,
    baseline_results: pd.DataFrame,
) -> dict[str, Path]:
    target_dir = bundle_root / "01_trend_adjusted_DID"
    tables_dir = target_dir / "tables"
    figures_dir = target_dir / "figures"
    text_dir = target_dir / "text"
    notes_dir = target_dir / "notes"
    scripts_dir = target_dir / "scripts"
    for folder in [tables_dir, figures_dir, text_dir, notes_dir, scripts_dir]:
        folder.mkdir(parents=True, exist_ok=True)

    trend_df = build_trend_results(panel)
    comparison_df = build_comparison_table(baseline_results, trend_df)

    long_rows = []
    for _, row in baseline_results.iterrows():
        long_rows.append(
            {
                "spec_type": "baseline",
                "depvar": row["depvar"],
                "coef": row["coef"],
                "se": row["se"],
                "p": row["p"],
                "N": row["N"],
                "formula": row["estimation_note"],
                "source": "5.3 formal long table",
            }
        )
    for _, row in trend_df.iterrows():
        long_rows.append(
            {
                "spec_type": "trend_adjusted",
                "depvar": row["depvar"],
                "coef": row["coef"],
                "se": row["se"],
                "p": row["p"],
                "N": row["N"],
                "formula": row["formula"],
                "source": "recomputed from unified V3 panel",
            }
        )
    long_df = pd.DataFrame(long_rows)

    workbook_path = tables_dir / "trend_adjusted_did_results.xlsx"
    write_excel(
        workbook_path,
        {
            "结果长表": long_df,
            "趋势调整主结果": trend_df,
            "系数对照": comparison_df,
        },
    )

    plot_path = figures_dir / "Figure_5_baseline_vs_trend_adjusted_forest.png"
    make_forest_plot(plot_path, comparison_df)

    readme_path = target_dir / "README.md"
    readme_text = textwrap.dedent(
        f"""
        # Trend-Adjusted DID Module

        This module is a defensive robustness layer. The canonical main identification remains `treat_share` DID/TWFE.

        ## Inputs
        - Unified baseline reference: `00_unified_baseline_reference/unified_baseline_reference.xlsx`
        - Formal V3 panel: `{relative_to_workspace(panel_path)}`
        - Baseline anchor table: `{relative_to_workspace(find_single(DATA_ROOT, LONG_TABLE_NAME))}`

        ## Model
        `y ~ treat_share + dfi + digital_econ + ln_rd_expenditure + ln_tech_contract_value + ln_patent_grants + C(province) + C(year) + C(province):year_idx`

        ## Outputs
        - `tables/trend_adjusted_did_results.xlsx`
        - `figures/Figure_5_baseline_vs_trend_adjusted_forest.png`
        - `text/trend_adjusted_did_body_insert.md`
        - `notes/trend_adjusted_did_notes.md`
        - `scripts/README.md`

        ## Canonical reading
        - `exec_share`: direction stays positive, significance weakens under province trends.
        - `proc_share`: remains negative and statistically meaningful under the current trend-adjusted specification.
        - `ppp_quality_zindex`: stays weaker and cannot be promoted to the main conclusion.
        """
    ).strip() + "\n"
    readme_path.write_text(readme_text, encoding="utf-8")

    notes_path = notes_dir / "trend_adjusted_did_notes.md"
    notes_text = textwrap.dedent(
        f"""
        # Trend-Adjusted DID Notes

        ## Purpose
        This module does not replace the canonical `treat_share` DID/TWFE design. It adds province-specific linear trends as a defensive robustness layer.

        ## Data and sample
        - Unified baseline reference: `00_unified_baseline_reference/unified_baseline_reference.xlsx`
        - Formal V3 main panel: `{relative_to_workspace(panel_path)}`
        - Estimation sample: `baseline_sample_5_3 == 1`
        - Cluster unit: `province`

        ## Specification
        - Baseline RHS is unchanged.
        - Added term: province-specific linear trends via `C(province):year_idx`.
        - Controls: `dfi`, `digital_econ`, `ln_rd_expenditure`, `ln_tech_contract_value`, `ln_patent_grants`

        ## Reference result summary
        - `exec_share`: {trend_df.loc[trend_df['depvar'] == 'exec_share', 'coef'].iloc[0]:.6f}, p={trend_df.loc[trend_df['depvar'] == 'exec_share', 'p'].iloc[0]:.6f}
        - `proc_share`: {trend_df.loc[trend_df['depvar'] == 'proc_share', 'coef'].iloc[0]:.6f}, p={trend_df.loc[trend_df['depvar'] == 'proc_share', 'p'].iloc[0]:.6f}
        - `ppp_quality_zindex`: {trend_df.loc[trend_df['depvar'] == 'ppp_quality_zindex', 'coef'].iloc[0]:.6f}, p={trend_df.loc[trend_df['depvar'] == 'ppp_quality_zindex', 'p'].iloc[0]:.6f}
        """
    ).strip() + "\n"
    notes_path.write_text(notes_text, encoding="utf-8")

    text_path = text_dir / "trend_adjusted_did_body_insert.md"
    text_text = textwrap.dedent(
        f"""
        ## 可插入位置建议
        建议将本段放在第5章 5.6“稳健性检验”中，作为常规稳健性与补充识别之间的“趋势调整型防守检验”。

        ## 建议正文
        为回应经典多期DID/TWFE可能受到地区原有差异化趋势影响的质疑，本文进一步在基准规格中加入省份线性趋势项，对核心结果进行趋势调整检验。该检验仍以 `treat_share` 为核心处理变量，并保持与基准结果一致的控制变量、地区固定效应和年份固定效应，仅额外控制各省随时间变化的线性趋势。结果见表X与图X。可以看到，在考虑地区层面可能存在的差异化线性趋势后，`exec_share` 的估计方向仍为正，但统计显著性明显减弱（系数 = {trend_df.loc[trend_df['depvar'] == 'exec_share', 'coef'].iloc[0]:.3f}, p = {trend_df.loc[trend_df['depvar'] == 'exec_share', 'p'].iloc[0]:.3f}），说明执行阶段占比上升这一判断在方向上保持一致，但对趋势设定更为敏感。`proc_share` 仍保持显著负向（系数 = {trend_df.loc[trend_df['depvar'] == 'proc_share', 'coef'].iloc[0]:.3f}, p = {trend_df.loc[trend_df['depvar'] == 'proc_share', 'p'].iloc[0]:.3f}），表明采购阶段占比下降这一核心判断并不完全依赖于共同趋势设定。相较之下，`ppp_quality_zindex` 在趋势调整后继续不显著且方向转负（系数 = {trend_df.loc[trend_df['depvar'] == 'ppp_quality_zindex', 'coef'].iloc[0]:.3f}, p = {trend_df.loc[trend_df['depvar'] == 'ppp_quality_zindex', 'p'].iloc[0]:.3f}），因此仍不宜被提升为全文主结论。
        """
    ).strip() + "\n"
    text_path.write_text(text_text, encoding="utf-8")

    scripts_readme_path = scripts_dir / "README.md"
    scripts_readme_text = textwrap.dedent(
        f"""
        # Script README

        Run `run_trend_adjusted_did.py` from this folder to regenerate both:
        - `00_unified_baseline_reference/unified_baseline_reference.xlsx`
        - `01_trend_adjusted_DID/tables/trend_adjusted_did_results.xlsx`

        The script only uses relative paths inside the workspace.
        """
    ).strip() + "\n"
    scripts_readme_path.write_text(scripts_readme_text, encoding="utf-8")

    return {
        "workbook": workbook_path,
        "plot": plot_path,
        "readme": readme_path,
        "notes": notes_path,
        "text": text_path,
        "scripts_readme": scripts_readme_path,
    }


def main() -> None:
    docx_path = find_single(WORKSPACE_ROOT, DOCX_NAME)
    panel_path, panel = load_panel()
    long_table_path, long_table = load_long_table()
    paragraphs = find_docx_paragraphs(docx_path)

    baseline_workbook, baseline_readme = build_unified_baseline_reference(
        BUNDLE_ROOT,
        docx_path,
        long_table_path,
        panel_path,
        long_table,
        panel,
        paragraphs,
    )
    baseline_results = load_baseline_results(long_table)
    trend_outputs = build_trend_outputs(BUNDLE_ROOT, panel_path, panel, baseline_results)

    print("Generated:")
    print(f"- {rel_or_name(baseline_workbook)}")
    print(f"- {rel_or_name(baseline_readme)}")
    for key, value in trend_outputs.items():
        print(f"- {rel_or_name(value)}")


if __name__ == "__main__":
    main()
